# -*- coding:utf-8 -*-

import openpyxl

from utils.conftest import args

"""引入命令行注册参数，像ls -help   python -v一样，把case绝对路径写入命令行，执行指定Excel"""
workbook = openpyxl.load_workbook(args.case)


def get_testcase():
    ExcelDataList = []
    sheets = workbook['Sheet1']
    rows_sheet = sheets.iter_rows()  # 一行一行取，如果用iter_cols()表示一列一列取
    for item in rows_sheet:
        if item[8].value == 'URI':
            continue
        new_list = []
        for col in item:
            new_list.append(col.value)
        ExcelDataList.append(new_list)
    return ExcelDataList


def get_testcase_line_no(case_no):
    # workbook = openpyxl.load_workbook(r'./testcase/接口自动化测试用例模板.xlsx')
    sheets = workbook['Sheet1']
    rows_sheet = sheets.iter_rows()  # 一行一行取，如果用iter_cols()表示一列一列取
    for item in rows_sheet:
        if item[0].value == '用例编号':
            continue
        for col in item:
            if case_no == col.value:
                row_no = col.row
    # print(row_no)
    return row_no


def write_result(line_no, column, excute_result):
    # workbook = openpyxl.load_workbook(r'./testcase/接口自动化测试用例模板.xlsx')
    sheets = workbook['Sheet1']
    sheets.cell(line_no, column, excute_result)
    workbook.save('./testcase/接口自动化测试用例模板.xlsx')


def get_line_no_testcase(line_no):
    case_no_info = get_testcase()[line_no - 2]
    return case_no_info


def get_case_info(case_no):
    line_no = get_testcase_line_no(case_no=case_no)  # 获取行号
    caseinfo = get_line_no_testcase(line_no=line_no)  # 获取当前case_no完整信息
    return caseinfo
